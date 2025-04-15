from django.shortcuts import render
from core.models import CartOrder, CartOrderItem, Product, Category, ProductReview, VendorOrder, Address, ProductImages
import datetime
from django.shortcuts import render, get_object_or_404
from Vendor.models import Vendor
from django.db.models import Sum
from userauths.models import Account, Profile, Contact, AdminRevenueRecord
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth.hashers import check_password, make_password
from django.utils.timezone import now
from decimal import Decimal
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date
import openpyxl
from core.utils import get_exchange_rate, convert_currency
from django.http import JsonResponse
from django.contrib.auth.models import Group
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q
from Vendor.forms import AddProductForm, ProductImageFormSet, CategoryForm
from django.db import models
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
import os
from dateutil.parser import parse
import uuid
import logging

logger = logging.getLogger(__name__)
# Create your views here.


@login_required
def AdminPanel(request):
    if not request.user.is_superadmin:
        return redirect('/custom_login/')
    
    revenue = AdminRevenueRecord.objects.first()
    total_orders_count = CartOrder.objects.all()
    all_products = Product.objects.all()
    all_categories = Category.objects.all()
    new_customers = Account.objects.all()
    latest_orders = CartOrder.objects.all()
    all_vendors = Vendor.objects.all()

    this_month = datetime.datetime.now().month
    monthly_revenue = CartOrder.objects.filter(order_date__month=this_month).aggregate(price=Sum("price"))

    context = {
        "revenue": revenue,
        "total_orders_count": total_orders_count,
        "all_products": all_products,
        "all_categories": all_categories,
        "all_vendors": all_vendors,
        "new_customers": new_customers,
        "latest_orders": latest_orders,
        "monthly_revenue": monthly_revenue,
    }
    return render(request, 'AdminPanel/AdminPanel.html', context)

@login_required
def products(request):
    if not request.user.is_superadmin:
        return redirect('/custom_login/')
    
    all_products = Product.objects.all().order_by("-id")
    all_categories = Category.objects.all()


    context = {
        "all_products": all_products,
    }

    return render(request, 'AdminPanel/products.html', context)



def addProduct(request):
    if request.method == 'POST':
        if not request.user.is_superadmin:
            return redirect('/custom_login/')
        
        form = AddProductForm(request.POST, request.FILES)
        formset = ProductImageFormSet(request.POST, request.FILES)

        if form.is_valid() and formset.is_valid():
            token = str(uuid.uuid4())

            if request.session.get('product_submission_token') == token:
                messages.error(request, "This product has already been added.")
                return redirect('addproduct')

            request.session['product_submission_token'] = token

            # Save the product
            product = form.save(commit=False)
            product.user = request.user
            product.product_status = 'published'
            product.featured = True
            product.currency = 'RWF'  # Ensure currency is RWF (redundant but safe)
            product.save()
            form.save_m2m()

            # Save the images
            for form in formset:
                if form.cleaned_data and form.cleaned_data.get('images'):
                    ProductImages.objects.create(product=product, images=form.cleaned_data['images'])
                if form.cleaned_data.get('DELETE'):
                    form.instance.delete()

            messages.success(request, "Product and images added successfully!")
            return redirect('addProduct')
        else:
            messages.error(request, "Error adding product. Please check the form.")
    else:
        form = AddProductForm()
        formset = ProductImageFormSet(queryset=ProductImages.objects.none())

    return render(request, 'AdminPanel/addProduct.html', {'form': form, 'formset': formset})




@login_required
def editProduct(request, pid):
    if not request.user.is_superadmin:
            return redirect('/custom_login/')
        
    product = get_object_or_404(Product, pid=pid)

    if request.method == 'POST':
        form = AddProductForm(request.POST, request.FILES, instance=product)
        formset = ProductImageFormSet(request.POST, request.FILES, queryset=ProductImages.objects.filter(product=product))

        if form.is_valid() and formset.is_valid():
            product = form.save(commit=False)
            product.product_status = 'published'
            product.featured = True
            product.currency = 'RWF'  # Explicitly ensure currency is RWF
            logger.debug(f"Saving product: price={product.price}, old_price={product.old_price}, currency={product.currency}")
            product.save()
            form.save_m2m()

            # Handle image updates/additions
            for form in formset:
                if form.cleaned_data:
                    image = form.save(commit=False)
                    image.product = product
                    image.save()

            messages.success(request, "Product and images updated successfully!")
            return redirect('products')
        else:
            logger.error(f"Form errors: {form.errors}")
            logger.error(f"Formset errors: {formset.errors}")
            messages.error(request, "Error updating product. Please check the form.")
    else:
        # Set initial currency to RWF to reflect stored value
        form = AddProductForm(instance=product, initial={'currency': 'RWF'})
        formset = ProductImageFormSet(queryset=ProductImages.objects.filter(product=product))

    return render(request, 'AdminPanel/editProduct.html', {
        'form': form,
        'formset': formset,
        'product': product
    })
    

@login_required
def deleteProduct(request, pid):
    if not request.user.is_superadmin:
            return redirect('/custom_login/')
        

    # Fetch the product by its ID and ensure it belongs to the current vendor (user)
    product = get_object_or_404(Product, pid=pid)

    # Delete the product
    product.delete()

    # Redirect to vendor products page after successful deletion
    return redirect("products")


@login_required
def products(request):
    if not request.user.is_superadmin:
        return render(request, 'userauths/custom_login.html')  
    
    # Get search query from GET parameters
    search_query = request.GET.get('q', '').strip()

    # Fetch all products and categories
    all_products = Product.objects.all().order_by("-id")
    if search_query:
        all_products = all_products.filter(title__icontains=search_query)

    all_categories = Category.objects.all()

    # Get the user's selected currency, default to RWF
    selected_currency = request.session.get('currency', 'RWF')

    # Fetch latest exchange rates with USD as base
    exchange_rates_usd = get_exchange_rate('USD')

    # Derive RWF-based exchange rates
    if exchange_rates_usd and 'RWF' in exchange_rates_usd:
        rwf_to_usd = Decimal('1') / Decimal(exchange_rates_usd['RWF'])
        exchange_rates_rwf = {k: Decimal(v) * rwf_to_usd for k, v in exchange_rates_usd.items()}
        exchange_rates_rwf['RWF'] = Decimal('1')
    else:
        exchange_rates_rwf = {
            'USD': Decimal('0.00074'),
            'EUR': Decimal('0.00069'),
            'GBP': Decimal('0.00057'),
            'RWF': Decimal('1'),
        }
        print("Using fallback RWF-based rates.")

    # Pagination
    paginator = Paginator(all_products, 10)  # 10 products per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Convert prices from RWF to selected currency
    for product in page_obj:
        product.converted_price = convert_currency(
            product.price, 'RWF', selected_currency, exchange_rates_rwf
        )
        product.converted_old_price = convert_currency(
            product.old_price, 'RWF', selected_currency, exchange_rates_rwf
        )

    context = {
        'page_obj': page_obj,
        'all_categories': all_categories,
        'selected_currency': selected_currency,
        'currency_symbols': {
            'RWF': 'RWF',
            'USD': '$',
            'EUR': '€',
            'GBP': '£',
        },
        'search_query': search_query,  # Pass search query to template
    }

    return render(request, 'AdminPanel/products.html', context)

@login_required
def delete_image(request, image_id):

    # Fetch the image by its ID and ensure it belongs to the current vendor
    image = get_object_or_404(ProductImages, id=image_id)

    # Delete the product
    image.delete()

    # Redirect to vendor products page after successful deletion
    return redirect("editProduct",pid=image.product.pid)



@login_required
def addCategory(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category added successfully.')
            return redirect('addCategory')
    else:
        form = CategoryForm()

    categories = Category.objects.all()
    return render(request, 'AdminPanel/addCategory.html', {'form': form, 'categories': categories})

@login_required
def editCategory(request, cid):
    category = get_object_or_404(Category, cid=cid)
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category updated successfully.')
            return redirect('addCategory')
    else:
        form = CategoryForm(instance=category)
    
    categories = Category.objects.all()
    return render(request, 'AdminPanel/addCategory.html', {'form': form, 'categories': categories, 'editing': category})

@login_required
def deleteCategory(request, cid):
    category = get_object_or_404(Category, cid=cid)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Category deleted successfully.')
        return redirect('addCategory')
    return redirect('addCategory')


@login_required
def orders(request):
    orders = CartOrder.objects.all()
    context = {
        "orders": orders
    }
    return render(request, 'AdminPanel/orders.html', context)



def is_admin(user):
    return user.is_superadmin

def createSeller(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        description = request.POST.get('description')
        address = request.POST.get('address')
        contact = request.POST.get('contact')
        image = request.FILES.get('image')
        id_image = request.FILES.get('id_image')
        is_seller = request.POST.get('is_seller') 
        username = email.split('@')[0]  

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect(request.path)

        if Account.objects.filter(email=email).exists():
            messages.error(request, "Email already exists.")
            return redirect(request.path)

        # Create user
        user = Account.objects.create(
            first_name=first_name,
            last_name=last_name,
            email=email,
            username=username,
            phone_number=contact,
            password=make_password(password),
            is_active=True
            
        )

        # Only make them a seller if checkbox is checked
        if is_seller == 'yes':
            sellers_group, created = Group.objects.get_or_create(name='Sellers')
            user.groups.add(sellers_group)

            Vendor.objects.create(
                user=user,
                image=image,
                id_image=id_image,
                description=description,
                address=address,
                contact=contact,
                email=email
            )

            messages.success(request, "Seller created and added to the sellers group.")
        else:
            messages.success(request, "User created without seller role.")

        return redirect('createSeller')

    return render(request, 'AdminPanel/createSeller.html')



@login_required
def vendor_list(request):
    all_vendors = Vendor.objects.all()
    for vendor in all_vendors:
        try:
            revenue_record = vendor.user.admin_revenue
            vendor.total_net_revenue = revenue_record.total_revenue
        except (AttributeError, AdminRevenueRecord.DoesNotExist):
            vendor.total_net_revenue = 0.00
    context = {
        "all_vendors": all_vendors
    }
    return render(request, 'AdminPanel/vendor_list.html', context)

@login_required
def vendor_detail(request, vid):
    vendor = get_object_or_404(Vendor, vid=vid)
    total_revenue = 0.00
    if vendor.user:
        try:
            total_revenue = vendor.user.admin_revenue.total_revenue
        except AdminRevenueRecord.DoesNotExist:
            pass  # Keep total_revenue as 0.00
    context = {
        'vendor': vendor,
        'total_revenue': total_revenue
    }
    return render(request, 'AdminPanel/vendor_detail.html', context)


@login_required
def order_details(request, oid):
    vendor_order = get_object_or_404(VendorOrder, id=oid)
    
    # Get the related CartOrder
    order = vendor_order.cart_order
    order_items = CartOrderItem.objects.filter(order=order)
    context = {
        'order': order,
        'order_items': order_items
    }

    return render(request, 'AdminPanel/order_details.html', context)



@login_required
def update_product_status(request, product_id):
    product = get_object_or_404(Product, id=product_id)  # Get the product

    if request.method == "POST":
        new_status = request.POST.get("product_status")  # Get new status from form
        product.product_status = new_status  # Update product status
        product.save()  # Save the changes
        messages.success(request, "Product status updated successfully!")  # Flash success message
        return redirect("products")  # Redirect to the previous page

    return redirect("dashboard")  # Redirect if not POST



@login_required
def customers(request):
    customers = Account.objects.filter(is_customer=True)
    
    context = {
        "customers":customers,
    }
    return render(request, 'AdminPanel/customers.html', context)


@login_required
def customer_detail_view(request, cid):
    customer = get_object_or_404(Account, id=cid, is_customer=True)
    customer_orders = CartOrder.objects.filter(customer=customer).order_by('-order_date')
    profile = Profile.objects.filter(user=customer).first()
    address = Address.objects.filter(user=customer).first()

    context = {
        'customer': customer,
        'customer_orders': customer_orders,
        'profile': profile,
        'address': address,
    }
    return render(request, 'AdminPanel/customer_detail_view.html', context)

@login_required
def change_paid_status(request, order_id):
    if request.method == "POST":
        order = get_object_or_404(CartOrder, id=order_id)
        paid_status = request.POST.get('paid_status')
        if paid_status not in ['True', 'False']:
            return HttpResponse("Invalid paid status", status=400)
        order.paid_status = (paid_status == 'True')
        order.save()

        if order.paid_status:
            admin_record, created = AdminRevenueRecord.objects.get_or_create(
                adminUser=request.user,  # Associate with the logged-in user
                defaults={'total_revenue': 0, 'monthly_revenue': 0}
            )
            admin_record.total_revenue += order.price or 0
            if order.order_date.month == timezone.now().month and order.order_date.year == timezone.now().year:
                admin_record.monthly_revenue += order.price or 0
            admin_record.save()

        return redirect('customer_detail_view', cid=order.customer.id)

    return HttpResponse("Invalid request", status=400)



@login_required
def single_order_detail(request, order_id):
    # Retrieve the order by ID (or return 404 if not found)
    order = get_object_or_404(CartOrder, id=order_id)
    
    # If you want to display customer information as well, you can get it from the `order.user`
    customer = order.user
    # Get the shipping address associated with the order (assuming you have it related)
    address = order.address  # or use a related model if necessary
    order_items = CartOrderItem.objects.filter(order=order)
    
    context = {
        'order': order,
        'customer': customer,
        'address': address,
        'order_items': order_items,
    }
    
    return render(request, 'AdminPanel/single_order_detail.html', context)


@login_required
def changePassword(request):
    user = request.user

    if request.method == "POST":
        old_password = request.POST.get("old_password")
        new_password = request.POST.get("new_password")
        confirm_new_password = request.POST.get("confirm_new_password")

        if confirm_new_password != new_password:
            messages.error(request, "Password Doesn't Match")
            return redirect("changePassword")
        
        if check_password(old_password, user.password):
            user.set_password(new_password)
            user.save()
            return redirect("userauths:custom_login")
        else:
            messages.error(request, "Old Password is Incorrect")
            return redirect("changePassword")
    return render(request, 'AdminPanel/changePassword.html' )

@login_required
def reviews(request):
    reviews = ProductReview.objects.all()

    context = {
        "reviews": reviews
    }
    return render(request, "AdminPanel/reviews.html", context)

@login_required
def Contacts(request):
    contacts = Contact.objects.all()

    context = {
        "contacts": contacts
    }
    return render(request, "AdminPanel/Contacts.html", context)


@login_required
def order_report(request):
    orders = CartOrder.objects.all()
    
    # Date Filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        orders = orders.filter(order_date__gte=parse_date(start_date))
    if end_date:
        orders = orders.filter(order_date__lte=parse_date(end_date))

    # Payment Method Filter
    payment_method = request.GET.get('payment_method')
    if payment_method:
        orders = orders.filter(payment_method=payment_method)  # Directly filter by payment_method field on CartOrder

    # Payment Status Filter
    payment_status = request.GET.get('payment_status')
    if payment_status == 'paid':
        orders = orders.filter(paid_status=True)  # Directly filter by paid_status field on CartOrder
    elif payment_status == 'not_paid':
        orders = orders.filter(paid_status=False)  # Directly filter by paid_status field on CartOrder

    # Billing Name Filter
    billing_name = request.GET.get('billing_name')
    if billing_name:
        orders = orders.filter(full_name__icontains=billing_name)  # Directly filter by full_name field on CartOrder
        
    context = {
        "orders": orders
    }
    return render(request, "AdminPanel/Orders.html", context)



@login_required
def export_order_report_excel(request):
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "MULTIVENDOR ECOMMERCE SYSTEM| Admin Order Report"

    # Define styles
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    cell_font = Font(name='Arial', size=10)
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Add Company Logo
    try:
        logo = Image('path/to/static/images/logo.png')  # Adjust path as needed
        logo.height = 50
        logo.width = 150
        ws.add_image(logo, 'A1')
    except:
        pass

    # Add Report Title and Date
    ws.merge_cells('A3:F3')
    ws['A3'] = "MULTIVENDOR ECOMMERCE SYSTEM| Admin Order Report"
    ws['A3'].font = title_font
    ws['A3'].alignment = center_alignment

    ws.merge_cells('A4:F4')
    ws['A4'] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'].font = Font(size=10, italic=True)
    ws['A4'].alignment = center_alignment

    # Get orders with the same filtering as order_report
    orders = CartOrder.objects.all()
    
    # Date Filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        orders = orders.filter(order_date__gte=parse(start_date))
    if end_date:
        orders = orders.filter(order_date__lte=parse(end_date))

    # Payment Method Filter
    payment_method = request.GET.get('payment_method')
    if payment_method:
        orders = orders.filter(payment_method=payment_method)

    # Payment Status Filter
    payment_status = request.GET.get('payment_status')
    if payment_status == 'paid':
        orders = orders.filter(paid_status=True)
    elif payment_status == 'not_paid':
        orders = orders.filter(paid_status=False)

    # Billing Name Filter
    billing_name = request.GET.get('billing_name')
    if billing_name:
        orders = orders.filter(full_name__icontains=billing_name)

    # Define headers (matching your template)
    headers = ['Order ID', 'Billing Name', 'Date', 'Total', 
              'Payment Method', 'Payment Status']
    ws.append([])  # Empty row for spacing
    ws.append(headers)

    # Style headers
    for col_num, value in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    # Add order data (using price as shown in your template)
    row_num = 7
    for order in orders:
        payment_status = 'Paid' if order.paid_status else 'Not Paid'
        ws.append([
            order.sku,  # Using sku as Order ID to match template
            order.full_name,
            order.order_date.strftime('%Y-%m-%d'),
            float(order.price) if order.price else 0.0,  # Using price as Total
            order.payment_method,
            payment_status
        ])
        
        # Style data rows
        for col_num in range(1, 7):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = cell_font
            cell.border = border
            if col_num in [1, 5, 6]:  # Left align for text
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif col_num == 4:  # Right align for amount
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
            else:  # Center align for others
                cell.alignment = center_alignment
        row_num += 1

    # Adjust column widths
    column_widths = [10, 25, 15, 15, 20, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Add totals row
    ws.append([])  # Empty row
    total_cell = ws.cell(row=row_num + 2, column=4)
    total_cell.value = f'=SUM(D7:D{row_num-1})'
    total_cell.font = header_font
    total_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_cell.number_format = '#,##0.00'
    total_cell.border = border

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Order_Report_{}.xlsx"'.format(
        timezone.now().strftime('%Y%m%d_%H%M%S')
    )

    wb.save(response)
    return response


@login_required
def vendor_report(request):
    # Extract filters from GET request
    vendor_status = request.GET.get('vendor_status', '')
    payment_status = request.GET.get('payment_status', '')
    search = request.GET.get('search', '')
    revenue = request.GET.get('revenue', '')

    # Initialize the vendor queryset
    all_vendors = Vendor.objects.all()

    # Apply filters based on the form inputs
    if vendor_status:
        all_vendors = all_vendors.filter(status=vendor_status)
        
    if payment_status:
        all_vendors = all_vendors.filter(payment_status=payment_status)

    if search:
        all_vendors = all_vendors.filter(
            Q(title__icontains=search) | 
            Q(contact__icontains=search) | 
            Q(email__icontains=search)
        )

    if revenue == 'highest_revenue':
        all_vendors = all_vendors.order_by('-total_net_amount')
    elif revenue == 'lowest_revenue':
        all_vendors = all_vendors.order_by('total_net_amount')
        
     # Assign total_net_revenue dynamically
    for vendor in all_vendors:
        vendor.total_net_revenue = vendor.total_net_amount or 0.00

    return render(request, 'AdminPanel/vendor_list.html', {
        'all_vendors': all_vendors,
        'vendor_status': vendor_status,
        'payment_status': payment_status,
        'search': search,
        'revenue': revenue,
    })



@login_required
def export_vendor_report_excel(request):
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "MULTIVENDOR ECOMMERCE SYSTEM| Admin Vendor Report"

    # Define styles
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    cell_font = Font(name='Arial', size=10)
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Add Company Logo
    try:
        logo = Image('path/to/static/images/logo.png')  # Adjust path as needed
        logo.height = 50
        logo.width = 150
        ws.add_image(logo, 'A1')
    except:
        pass

    # Add Report Title and Date
    ws.merge_cells('A3:H3')
    ws['A3'] = "MULTIVENDOR ECOMMERCE SYSTEM| Admin Vendor Report"
    ws['A3'].font = title_font
    ws['A3'].alignment = center_alignment

    ws.merge_cells('A4:H4')
    ws['A4'] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'].font = Font(size=10, italic=True)
    ws['A4'].alignment = center_alignment

    # Get vendors with the same filtering as vendor_report
    all_vendors = Vendor.objects.all()
    
    # Extract filters from GET request
    vendor_status = request.GET.get('vendor_status', '')
    payment_status = request.GET.get('payment_status', '')
    search = request.GET.get('search', '')
    revenue = request.GET.get('revenue', '')

    # Apply filters
    if vendor_status:
        all_vendors = all_vendors.filter(status=vendor_status)
        
    if payment_status:
        all_vendors = all_vendors.filter(payment_status=payment_status)

    if search:
        all_vendors = all_vendors.filter(
            Q(title__icontains=search) | 
            Q(contact__icontains=search) | 
            Q(email__icontains=search)
        )

    if revenue == 'highest_revenue':
        all_vendors = all_vendors.order_by('-total_net_amount')
    elif revenue == 'lowest_revenue':
        all_vendors = all_vendors.order_by('total_net_amount')

    # Define headers (matching your template)
    headers = ['Vendor ID', 'Title', 'Contact', 'Subscription', 'Status', 'Revenue']
    ws.append([])  # Empty row for spacing
    ws.append(headers)

    # Style headers
    for col_num, value in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    # Add vendor data
    row_num = 7
    for vendor in all_vendors:
        subscription_status = 'Paid' if vendor.has_paid_fee else 'Not Paid'
        vendor_status = vendor.get_status_display()  # Use display value as in template
        total_net_revenue = float(vendor.total_net_amount or 0.00)  # Match total_net_revenue logic
        ws.append([
            vendor.vid,
            vendor.title,
            vendor.contact,
            subscription_status,
            vendor_status,
            total_net_revenue
        ])
        
        # Style data rows
        for col_num in range(1, 7):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = cell_font
            cell.border = border
            if col_num in [2, 3]:  # Left align for Title and Contact
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif col_num == 6:  # Right align for Revenue
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
            else:  # Center align for others
                cell.alignment = center_alignment
        row_num += 1

    # Adjust column widths
    column_widths = [15, 25, 20, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Add totals row for Revenue
    ws.append([])  # Empty row
    total_cell = ws.cell(row=row_num + 2, column=6)
    total_cell.value = f'=SUM(F7:F{row_num-1})'
    total_cell.font = header_font
    total_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_cell.number_format = '#,##0.00'
    total_cell.border = border

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Vendor_Report_{}.xlsx"'.format(
        timezone.now().strftime('%Y%m%d_%H%M%S')
    )

    wb.save(response)
    return response







@login_required
def customer_report(request):
    customers = Account.objects.filter(is_customer=True)

    # Get filter values from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    search_query = request.GET.get('search')

    # Filter by date range
    if start_date:
        customers = customers.filter(date_joined__gte=parse_date(start_date))
    if end_date:
        customers = customers.filter(date_joined__lte=parse_date(end_date))

    # Search by name, email, contact, or username
    if search_query:
        customers = customers.filter(
            models.Q(username__icontains=search_query) |
            models.Q(first_name__icontains=search_query) |
            models.Q(last_name__icontains=search_query) |
            models.Q(email__icontains=search_query) |
            models.Q(phone_number__icontains=search_query)
        )

    context = {
        'customers': customers
    }
    return render(request, 'AdminPanel/customers.html', context)



@login_required
def export_customer_report_excel(request):
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "MULTIVENDOR ECOMMERCE SYSTEM| Customer Report"

    # Define styles
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    cell_font = Font(name='Arial', size=10)
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Add Company Logo
    try:
        logo = Image('path/to/static/images/logo.png')  # Adjust path as needed
        logo.height = 50
        logo.width = 150
        ws.add_image(logo, 'A1')
    except:
        pass

    # Add Report Title and Date
    ws.merge_cells('A3:G3')
    ws['A3'] = "MULTIVENDOR ECOMMERCE SYSTEM| Customer Report"
    ws['A3'].font = title_font
    ws['A3'].alignment = center_alignment

    ws.merge_cells('A4:G4')
    ws['A4'] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'].font = Font(size=10, italic=True)
    ws['A4'].alignment = center_alignment

    # Get customers with the same filtering as customer_report
    customers = Account.objects.filter(is_customer=True)
    
    # Get filter values from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    search_query = request.GET.get('search')

    # Filter by date range
    if start_date:
        customers = customers.filter(date_joined__gte=parse(start_date))
    if end_date:
        customers = customers.filter(date_joined__lte=parse(end_date))

    # Search by name, email, contact, or username
    if search_query:
        customers = customers.filter(
            models.Q(username__icontains=search_query) |
            models.Q(first_name__icontains=search_query) |
            models.Q(last_name__icontains=search_query) |
            models.Q(email__icontains=search_query) |
            models.Q(phone_number__icontains=search_query)
        )

    # Define headers (matching your template)
    headers = ['Joined Date', 'Username', 'Full Name', 'Contact', 'Email', 'Status']
    ws.append([])  # Empty row for spacing
    ws.append(headers)

    # Style headers
    for col_num, value in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    # Add customer data
    row_num = 7
    for customer in customers:
        full_name = f"{customer.first_name} {customer.last_name}".strip()
        status = 'Active' if customer.is_active else 'Not Active'
        ws.append([
            customer.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
            customer.username,
            full_name,
            customer.phone_number,
            customer.email,
            status
        ])
        
        # Style data rows
        for col_num in range(1, 7):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = cell_font
            cell.border = border
            if col_num in [2, 3, 5]:  # Left align for Full Name, Contact, Email
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # Center align for others
                cell.alignment = center_alignment
        row_num += 1

    # Adjust column widths
    column_widths = [20, 15, 25, 15, 25, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Customer_Report_{}.xlsx"'.format(
        timezone.now().strftime('%Y%m%d_%H%M%S')
    )

    wb.save(response)
    return response


