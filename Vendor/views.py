from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils.dateparse import parse_date
from .forms import CategoryForm
from openpyxl import Workbook
from django.utils import timezone
from django.core.paginator import Paginator
from django.contrib.auth.hashers import check_password
from core.models import Category, Product, CartOrder, VendorOrder, CartOrderItem
from django.db.models import Sum, Count
from datetime import datetime
from .models import Vendor
from core.utils import get_exchange_rate, convert_currency
from decimal import Decimal
from .utils import get_vendor_status
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
import openpyxl
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.parser import parse
import logging
from userauths.models import Account
import re
from django.db import transaction, IntegrityError
import uuid
logger = logging.getLogger(__name__)




@login_required
def vendorDashboard(request):
    vendor, redirect_response = get_vendor_status(request)
    if redirect_response:  # If the helper returned a redirect or render response, use it
        return redirect_response
    
    # Fetch VendorOrders related to this vendor where the related CartOrder's payment status is 'Paid'
    vendor_orders = VendorOrder.objects.filter(vendor=vendor, cart_order__paid_status=True)

    

    # Count the number of vendor orders and products
    vendor_order_count = vendor_orders.count()
    product_count = Product.objects.all().count()

    # Get current month and year
    current_month = datetime.now().month
    current_year = datetime.now().year

    # Calculate monthly earnings only
    monthly_earnings = vendor_orders.filter(order_date__month=current_month, order_date__year=current_year)\
                                    .aggregate(total=Sum('net_amount'))['total'] or 0

    # Get the user's selected currency (default to USD if not set)
    selected_currency = request.session.get('currency', 'USD')

    # Fetch exchange rates and convert earnings
    exchange_rates = get_exchange_rate('USD')
    
    converted_monthly_earnings = convert_currency(monthly_earnings, 'USD', selected_currency, exchange_rates)

    # Convert each vendor order's amounts
    for order in vendor_orders:
        order.converted_total_amount = convert_currency(order.total_amount, 'USD', selected_currency, exchange_rates)
        order.converted_commission = convert_currency(order.commission, 'USD', selected_currency, exchange_rates)
        order.converted_net_amount = convert_currency(order.net_amount, 'USD', selected_currency, exchange_rates)

    context = {
        'vendor': vendor,
        'vendor_orders': vendor_orders,
        'vendor_order_count': vendor_order_count,
        'product_count': product_count,
        'monthly_earnings': monthly_earnings,
        'selected_currency': selected_currency,
        'converted_monthly_earnings': converted_monthly_earnings,
    }

    # Render the dashboard
    return render(request, 'Vendor/vendorDashboard.html', context)




@login_required
def vendorProducts(request):
    vendor, redirect_response = get_vendor_status(request)
    if redirect_response:  # If the helper returned a redirect or render response, use it
        return redirect_response  
    
    # Get search query and category filter from GET parameters
    search_query = request.GET.get('q', '').strip()
    category_filter = request.GET.get('category', '').strip()

    # Fetch all products and categories
    all_products = Product.objects.all().order_by("-id")
    if search_query or category_filter:
        query = Q()
        if search_query:
            query |= Q(title__icontains=search_query) | Q(category__title__icontains=search_query)
        if category_filter:
            query &= Q(category__id=category_filter)
        all_products = all_products.filter(query)

    all_categories = Category.objects.all()

    # Get the user's selected currency, default to RWF
    selected_currency = request.session.get('currency', 'RWF')

    # Fetch latest exchange rates with USD as base
    exchange_rates_usd = get_exchange_rate('USD')

    # Derive RWF-based exchange rates
    if exchange_rates_usd and 'RWF' in exchange_rates_usd:
        rwf_to_usd = Decimal('1') / Decimal(exchange_rates_usd['RWF'])
        exchange_rates_rwf = {k: Decimal(v) * rwf_to_usd for k, v in exchange_rates_usd.items()}
        exchange_rates_rwf['RWF'] = Decimal('1')
    else:
        exchange_rates_rwf = {
            'USD': Decimal('0.00074'),
            'EUR': Decimal('0.00069'),
            'GBP': Decimal('0.00057'),
            'RWF': Decimal('1'),
        }
        print("Using fallback RWF-based rates.")

    # Pagination
    paginator = Paginator(all_products, 10)  # 10 products per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Convert prices from RWF to selected currency
    for product in page_obj:
        product.converted_price = convert_currency(
            product.price, 'RWF', selected_currency, exchange_rates_rwf
        )
        product.converted_old_price = convert_currency(
            product.old_price, 'RWF', selected_currency, exchange_rates_rwf
        )

    context = {
        'page_obj': page_obj,
        'all_categories': all_categories,
        'selected_currency': selected_currency,
        'currency_symbols': {
            'RWF': 'RWF',
            'USD': '$',
            'EUR': '€',
            'GBP': '£',
        },
        'search_query': search_query,
        'category_filter': category_filter,
    }

    return render(request, 'Vendor/vendorProducts.html', context)







@login_required
def vendorOrders(request):
    # Use the helper function to check vendor status and handle redirects
    vendor, redirect_response = get_vendor_status(request)
    if redirect_response:  # If the helper returned a redirect or render response, use it
        return redirect_response

    # Fetch VendorOrders related to this vendor
    vendor_orders = VendorOrder.objects.filter(vendor=vendor)
    selected_currency = request.session.get('currency', 'USD')  # Default to USD if not set

    # Fetch the latest exchange rates (base currency: USD)
    exchange_rates = get_exchange_rate('USD')  # Fetch exchange rates for USD base

    # Convert each vendor order amounts
    for order in vendor_orders:
        order.converted_total_amount = convert_currency(order.total_amount, 'USD', selected_currency, exchange_rates)
        order.converted_commission = convert_currency(order.commission, 'USD', selected_currency, exchange_rates)
        order.converted_net_amount = convert_currency(order.net_amount, 'USD', selected_currency, exchange_rates)

    # Extract CartOrders from VendorOrders
    cart_orders = CartOrder.objects.filter(id__in=vendor_orders.values('cart_order'))

    context = {
        'vendor': vendor,
        'vendor_orders': vendor_orders,
        'cart_orders': cart_orders,
        'selected_currency': selected_currency,
    }

    # Render the vendor orders page
    return render(request, 'Vendor/vendorOrders.html', context)



def vendor_orders_view(request):
    
    vendor, redirect_response = get_vendor_status(request)
    vendor_orders = VendorOrder.objects.filter(vendor=request.user.vendor)

    # Date Filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        vendor_orders = vendor_orders.filter(order_date__gte=parse_date(start_date))
    if end_date:
        vendor_orders = vendor_orders.filter(order_date__lte=parse_date(end_date))

    # Payment Method Filter
    payment_method = request.GET.get('payment_method')
    if payment_method:
        vendor_orders = vendor_orders.filter(cart_order__payment_method=payment_method)

    # Payment Status Filter
    payment_status = request.GET.get('payment_status')
    if payment_status == 'paid':
        vendor_orders = vendor_orders.filter(cart_order__paid_status=True)
    elif payment_status == 'not_paid':
        vendor_orders = vendor_orders.filter(cart_order__paid_status=False)

    billing_name = request.GET.get('billing_name')
    if billing_name:
        vendor_orders = vendor_orders.filter(cart_order__full_name__icontains=billing_name)
    
    
    selected_currency = request.session.get('currency', 'USD')  # Default to USD if not set

    # Fetch the latest exchange rates (base currency: USD)
    exchange_rates = get_exchange_rate('USD')  # Fetch exchange rates for USD base

    # Convert each vendor order amounts
    for order in vendor_orders:
        order.converted_total_amount = convert_currency(order.total_amount, 'USD', selected_currency, exchange_rates)
        order.converted_commission = convert_currency(order.commission, 'USD', selected_currency, exchange_rates)
        order.converted_net_amount = convert_currency(order.net_amount, 'USD', selected_currency, exchange_rates)
        
    context = {
        'vendor_orders': vendor_orders,
    }
    return render(request, 'Vendor/vendorOrders.html', context)




@login_required
def export_vendor_orders_excel(request):
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "MULTIVENDOR ECOMMERCE SYSTEM| Vendor Orders Report"

    # Define styles
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    cell_font = Font(name='Arial', size=10)
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Add Company Logo
    try:
        logo = Image('path/to/static/images/logo.png')  # Adjust path as needed
        logo.height = 50
        logo.width = 150
        ws.add_image(logo, 'A1')
    except:
        pass

    # Add Report Title and Date
    ws.merge_cells('A3:I3')
    ws['A3'] = "MULTIVENDOR ECOMMERCE SYSTEM| Vendor Orders Report"
    ws['A3'].font = title_font
    ws['A3'].alignment = center_alignment

    ws.merge_cells('A4:I4')
    ws['A4'] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'].font = Font(size=10, italic=True)
    ws['A4'].alignment = center_alignment

    # Get vendor and orders with the same filtering as vendor_orders_view
    vendor, redirect_response = get_vendor_status(request)
    if redirect_response:
        return redirect_response  # Handle redirect if vendor status check fails
    
    vendor_orders = VendorOrder.objects.filter(vendor=request.user.vendor)
    
    # Date Filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        vendor_orders = vendor_orders.filter(order_date__gte=parse(start_date))
    if end_date:
        vendor_orders = vendor_orders.filter(order_date__lte=parse(end_date))

    # Payment Method Filter
    payment_method = request.GET.get('payment_method')
    if payment_method:
        vendor_orders = vendor_orders.filter(cart_order__payment_method=payment_method)

    # Payment Status Filter
    payment_status = request.GET.get('payment_status')
    if payment_status == 'paid':
        vendor_orders = vendor_orders.filter(cart_order__paid_status=True)
    elif payment_status == 'not_paid':
        vendor_orders = vendor_orders.filter(cart_order__paid_status=False)

    # Billing Name Filter
    billing_name = request.GET.get('billing_name')
    if billing_name:
        vendor_orders = vendor_orders.filter(cart_order__full_name__icontains=billing_name)

    # Currency conversion (same as in view)
    selected_currency = request.session.get('currency', 'USD')  # Default to USD
    exchange_rates = get_exchange_rate('USD')  # Fetch exchange rates for USD base

    # Define headers (matching your template)
    headers = ['Order ID', 'Billing Name', 'Date', 'Total', 'Commission', 'Net Amount', 
              'Payment Method', 'Payment Status']
    ws.append([])  # Empty row for spacing
    ws.append(headers)

    # Style headers
    for col_num, value in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    # Add vendor order data with converted amounts
    row_num = 7
    for order in vendor_orders:
        # Convert amounts
        converted_total_amount = convert_currency(order.total_amount, 'USD', selected_currency, exchange_rates)
        converted_commission = convert_currency(order.commission, 'USD', selected_currency, exchange_rates)
        converted_net_amount = convert_currency(order.net_amount, 'USD', selected_currency, exchange_rates)
        
        payment_method_display = 'Cash On Delivery' if order.cart_order.payment_method == 'cod' else order.cart_order.payment_method
        payment_status = 'Paid' if order.cart_order.paid_status else 'Not Paid'
        
        ws.append([
            order.cart_order.sku,
            order.cart_order.full_name,
            order.order_date.strftime('%Y-%m-%d'),
            float(converted_total_amount) if converted_total_amount else 0.0,
            float(converted_commission) if converted_commission else 0.0,
            float(converted_net_amount) if converted_net_amount else 0.0,
            payment_method_display,
            payment_status
        ])
        
        # Style data rows
        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = cell_font
            cell.border = border
            if col_num in [1, 7]:  # Left align for Billing Name and Payment Method
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif col_num in [4, 5, 6]:  # Right align for Total, Commission, Net Amount
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
            else:  # Center align for others
                cell.alignment = center_alignment
        row_num += 1

    # Adjust column widths
    column_widths = [15, 25, 15, 15, 15, 15, 20, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Add totals row for Total, Commission, and Net Amount
    ws.append([])  # Empty row
    for col, formula in [(4, 'D'), (5, 'E'), (6, 'F')]:  # Total, Commission, Net Amount
        total_cell = ws.cell(row=row_num + 2, column=col)
        total_cell.value = f'=SUM({formula}7:{formula}{row_num-1})'
        total_cell.font = header_font
        total_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_cell.number_format = '#,##0.00'
        total_cell.border = border

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Vendor_Orders_Report_{selected_currency}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response




@login_required
def vendor_order_Details(request, vendor_order_id):
    # Get the VendorOrder instance
    vendor_order = get_object_or_404(VendorOrder, id=vendor_order_id)
    
    # Get the related CartOrder
    cart_order = vendor_order.cart_order
    
    # Get the CartOrderItem instances related to the CartOrder
    cart_order_items = CartOrderItem.objects.filter(order=cart_order)

    selected_currency = request.session.get('currency', 'USD') 
    exchange_rates = get_exchange_rate('USD') 

    cart_order_total_converted = convert_currency(cart_order.price, 'USD', selected_currency, exchange_rates)
    saved_converted = convert_currency(cart_order.saved, 'USD', selected_currency, exchange_rates)
    shipping_fee = convert_currency(cart_order.shipping_fee, 'USD', selected_currency, exchange_rates)
    final_price = convert_currency(cart_order.final_price, 'USD', selected_currency,exchange_rates)

    for item in cart_order_items:
        item.converted_price = convert_currency(item.product.price, 'USD', selected_currency, exchange_rates)
        item.converted_total = convert_currency(item.total, 'USD', selected_currency, exchange_rates)

    # Handle form submission
    if request.method == "POST":
        new_status = request.POST.get('order_status')
        if new_status in ['processing','out_for_delivery', 'delivered']:
            vendor_order.cart_order.product_status = new_status
            vendor_order.cart_order.save()
            messages.success(request, f"Order status updated to {new_status}.")
        return redirect('vendor_order_Details', vendor_order_id=vendor_order_id)

    context = {
        'vendor_order': vendor_order,
        'cart_order': cart_order,
        'cart_order_items': cart_order_items,
        'cart_order_total_converted': cart_order_total_converted,
        'saved_converted': saved_converted,
        'shipping_fee': shipping_fee,
        'selected_currency': selected_currency,
        'final_price':final_price,
    }
    
    return render(request, 'Vendor/vendor_order_Details.html', context)


@login_required
def vendorStore(request):
    # Use the helper function to check vendor status and handle redirects
    vendor, redirect_response = get_vendor_status(request)
    if redirect_response:  # If the helper returned a redirect or render response, use it
        return redirect_response

    # Fetch products related to the vendor
    products = Product.objects.all()

    # Fetch VendorOrders related to this vendor where the related CartOrder's payment status is 'Paid'
    vendor_orders = VendorOrder.objects.filter(vendor=vendor, cart_order__paid_status=True)

    

    selected_currency = request.session.get('currency', 'USD')  # Default to USD if not set

    # Fetch the latest exchange rates (base currency: USD)
    exchange_rates = get_exchange_rate('USD')

    for product in products:
        product.converted_price = convert_currency(product.price, 'USD', selected_currency, exchange_rates)

    # Count the number of vendor orders
    vendor_order_count = vendor_orders.aggregate(Count('id'))['id__count']

    # Pagination
    paginator = Paginator(products, 10)  # Show 10 products per page
    page_number = request.GET.get('page')
    paginated_products = paginator.get_page(page_number)

    # Render the vendor store page with products
    context = {
        'vendor': vendor,
        'products': paginated_products,
        'vendor_order_count': vendor_order_count,
        'selected_currency': selected_currency,
        'paginator': paginator,  # Optional: include paginator for additional controls
    }
    return render(request, 'Vendor/vendorStore.html', context)


def InStoreOrder(request):
    if not request.user.groups.filter(name='Sellers').exists():
        messages.error(request, 'You do not have permission to access the vendor dashboard.')
        return redirect('Home')

    vendor, redirect_response = get_vendor_status(request)
    if redirect_response:
        return redirect_response

    products = Product.objects.filter(product_status='published')
    categories = Category.objects.all().order_by('level', 'title')
    form_data = {}
    errors = {}

    if request.method == 'POST':
        customer_email = request.POST.get('customer_email', '').strip()
        customer_first_name = request.POST.get('customer_first_name', '').strip()
        customer_last_name = request.POST.get('customer_last_name', '').strip()
        customer_phone = request.POST.get('customer_phone', '').strip()
        address = request.POST.get('address', '').strip()
        city = request.POST.get('city', '').strip()
        country = request.POST.get('country', '').strip()
        payment_method = request.POST.get('payment_method', '').strip()
        selected_currency = request.POST.get('currency', 'RWF')
        product_ids = request.POST.getlist('product')
        quantities = request.POST.getlist('quantity')

        form_data = {
            'customer_email': customer_email,
            'customer_first_name': customer_first_name,
            'customer_last_name': customer_last_name,
            'customer_phone': customer_phone,
            'address': address,
            'city': city,
            'country': country,
            'payment_method': payment_method,
            'currency': selected_currency,
            'items': [
                {'pid': pid, 'quantity': qty}
                for pid, qty in zip(product_ids, quantities)
            ]
        }

        if customer_email and not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', customer_email):
            messages.error(request, "Please enter a valid email address.")
            errors['customer_email'] = "Invalid email format."
        if customer_first_name and not re.match(r'^[a-zA-Z\s-]+$', customer_first_name):
            messages.error(request, "First name can only contain letters, spaces, or hyphens.")
            errors['customer_first_name'] = "Invalid first name format."
        if customer_last_name and not re.match(r'^[a-zA-Z\s-]+$', customer_last_name):
            messages.error(request, "Last name can only contain letters, spaces, or hyphens.")
            errors['customer_last_name'] = "Invalid last name format."
        if customer_phone and not re.match(r'^\+2507\d{8}$', customer_phone):
            messages.error(request, "Phone number must be exactly 12 digits starting with +2507.")
            errors['customer_phone'] = "Invalid phone format."
        if address and not re.match(r'^[a-zA-Z0-9\s,.-]+$', address):
            messages.error(request, "Address can only contain letters, numbers, spaces, commas, periods, or hyphens.")
            errors['address'] = "Invalid address format."
        if city and not re.match(r'^[a-zA-Z\s-]+$', city):
            messages.error(request, "City can only contain letters, spaces, or hyphens.")
            errors['city'] = "Invalid city format."
        if country and not re.match(r'^[a-zA-Z\s-]+$', country):
            messages.error(request, "Country can only contain letters, spaces, or hyphens.")
            errors['country'] = "Invalid country format."
        if payment_method and payment_method not in ['cash', 'momo', 'airtel']:
            messages.error(request, "Invalid payment method selected.")
            errors['payment_method'] = "Please select a valid payment method."

        if not product_ids or not quantities or len(product_ids) != len(quantities):
            messages.error(request, "Please add at least one valid product and quantity.")
            errors['items'] = "Invalid product or quantity selection."
        elif len(product_ids) != len(set(product_ids)):
            messages.error(request, "Duplicate products selected. Please combine quantities.")
            errors['items'] = "Duplicate products are not allowed."
        else:
            item_errors = []
            for idx, (pid, qty) in enumerate(zip(product_ids, quantities)):
                try:
                    qty = int(qty)
                    if qty < 1:
                        item_errors.append(f"Quantity for item {idx + 1} must be at least 1.")
                        continue
                    product = Product.objects.get(pid=pid, product_status='published')
                    if product.in_stock < qty:
                        item_errors.append(f"Not enough stock for {product.title} ({product.in_stock} available).")
                except ValueError:
                    item_errors.append(f"Invalid quantity for item {idx + 1}.")
                except Product.DoesNotExist:
                    item_errors.append(f"Product with ID {pid} does not exist.")
            if item_errors:
                errors['items'] = item_errors
                for error in item_errors:
                    messages.error(request, error)

        if not errors:
            try:
                with transaction.atomic():
                    exchange_rates = get_exchange_rate('USD')

                    customer = None
                    if customer_email:
                        try:
                            customer = Account.objects.get(email=customer_email, is_customer=True)
                            if customer_first_name and not customer.first_name:
                                customer.first_name = customer_first_name
                            if customer_last_name and not customer.last_name:
                                customer.last_name = customer_last_name
                            if customer_phone and not customer.phone_number:
                                customer.phone_number = customer_phone
                            customer.save()
                        except Account.DoesNotExist:
                            try:
                                customer = Account.objects.create(
                                    email=customer_email,
                                    username=customer_email.split('@')[0],
                                    is_customer=True,
                                    first_name=customer_first_name,
                                    last_name=customer_last_name,
                                    phone_number=customer_phone
                                )
                            except IntegrityError as e:
                                logger.error(f"Failed to create customer: {e}")
                                messages.error(request, "Invalid customer email or phone number.")
                                raise ValueError("Customer creation failed.")

                    order = CartOrder.objects.create(
                        user=request.user,
                        customer=customer,
                        full_name=f"{customer_first_name} {customer_last_name}".strip() if customer_first_name or customer_last_name else None,
                        email=customer_email or None,
                        phone_number=customer_phone or None,
                        address=address or None,
                        city=city or None,
                        country=country or None,
                        payment_method=payment_method or None,
                        shipping_fee=0,
                        traching_id=str(uuid.uuid4())[:8],
                        paid_status=payment_method == 'cash',
                        order_type='in_store',
                        currency='RWF',
                        price=0,
                        sku=str(uuid.uuid4())[:8],
                        oid=str(uuid.uuid4())[:8],
                        order_date=timezone.now()
                    )

                    total_rwf = Decimal('0.00')
                    for pid, qty in zip(product_ids, quantities):
                        try:
                            qty = int(qty)
                            product = Product.objects.get(pid=pid, product_status='published')
                            image_url = product.image.url if product.image else ''

                            price_rwf = convert_currency(product.price, product.currency, 'RWF', exchange_rates)
                            if price_rwf is None:
                                raise ValueError(f"Failed to convert {product.price} {product.currency} to RWF")
                            
                            total_item_rwf = price_rwf * Decimal(qty)

                            CartOrderItem.objects.create(
                                order=order,
                                product=product,
                                invoice_no=str(uuid.uuid4())[:8],
                                product_status='processing',
                                item=product.title,
                                image=image_url,
                                qty=qty,
                                price=price_rwf,
                                total=total_item_rwf,
                            )

                            product.in_stock -= qty
                            product.save()

                            total_rwf += total_item_rwf

                        except ValueError as ve:
                            logger.error(f"Error processing item {pid}: {ve}")
                            messages.error(request, f"Error processing item: {str(ve)}")
                            order.delete()
                            raise
                        except Product.DoesNotExist:
                            logger.error(f"Product with pid {pid} not found.")
                            messages.error(request, f"Product with ID {pid} does not exist.")
                            order.delete()
                            raise
                        except Exception as e:
                            logger.error(f"Unexpected error for item {pid}: {e}")
                            messages.error(request, f"Error processing item: {str(e)}")
                            order.delete()
                            raise

                    order.price = total_rwf
                    order.final_price = total_rwf
                    order.save()
                    logger.info(f"Order {order.oid} created with total {total_rwf} RWF")
                    messages.success(request, "In-store order created successfully!")
                    return redirect('vendorDashboard')

            except Exception as e:
                logger.error(f"Order processing failed: {e}")
                messages.error(request, f"Failed to process order: {str(e)}")
                errors['items'] = str(e)

        return render(request, 'Vendor/InStoreOrder.html', {
            'products': products,
            'categories': categories,
            'form_data': form_data,
            'errors': errors,
        })

    return render(request, 'Vendor/InStoreOrder.html', {
        'products': products,
        'categories': categories,
        'form_data': {'items': [{}]},
        'errors': {},
    })

@login_required
def vendorAddCategory(request):
    vendor, redirect_response = get_vendor_status(request)
    if redirect_response:  # If the helper returned a redirect or render response, use it
        return redirect_response

    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category added successfully.')
            return redirect('vendorAddCategory')  # Redirect to a page where categories are listed
    else:
        form = CategoryForm()

    categories = Category.objects.all()  # Fetch all categories

    return render(request, 'Vendor/vendorAddCategory.html', {'form': form, 'categories': categories})








@login_required 
def vendorChangepswd(request):
    user = request.user

    if request.method == "POST":
        old_password = request.POST.get("old_password")
        new_password = request.POST.get("new_password")
        confirm_new_password = request.POST.get("confirm_new_password")

        if confirm_new_password != new_password:
            messages.error(request, "Password Doesn't Match")
            return redirect("change_password")
        
        if check_password(old_password, user.password):
            user.set_password(new_password)
            user.save()
            return redirect("userauths:custom_login")
        else:
            messages.error(request, "Old Password is Incorrect")
            return redirect("change_password")
        
    return render(request, "Vendor/vendorChangepswd.html")




    
    
from django.db.models import Q


def search_vendors(query):
    if not query:
        return Vendor.objects.all()

    vendors = Vendor.objects.filter(
        Q(title__icontains=query) |
        Q(email__icontains=query) |
        Q(status__icontains=query) |
        Q(contact__icontains=query)
    ).distinct()

    return vendors

# In your view
from django.shortcuts import render

def search_vendor(request):
    query = request.GET.get('q', '').strip()
    vendors = search_vendors(query)
    return render(request, 'Core/search_vendor.html', {'vendors': vendors, 'query': query})


def vendor_product(request):
    # Use the helper function to check vendor status and handle redirects
    vendor, redirect_response = get_vendor_status(request)
    if redirect_response:  # If the helper returned a redirect or render response, use it
        return redirect_response
    
    product = Product.objects.filter(vendor=vendor)
    
    return render(request, 'Vendor/vendor_product.html')